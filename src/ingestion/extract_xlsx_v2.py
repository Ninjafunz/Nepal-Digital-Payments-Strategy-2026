"""
CORRECT EXTRACTION: NRB Payment Systems Indicators XLSX → CSV

Known issues in the previous extraction (v1):
1. FY total columns in Sections B1/B2 were treated as monthly data
2. This caused cascading 1-month shifts after each FY boundary  
3. Final months were dropped

This script uses Section A (no FY totals) as canonical month list,
then skips FY total columns when reading B1/B2.
"""
import openpyxl
import csv
import os

XLSX_CANDIDATES = [
    'sources/NRB_PSD_Sep2025_REAL.xlsx',
]
OUTPUT_PATH = 'data/raw/NRB_Payment_Systems_Indicators_monthly.csv'

def resolve_xlsx():
    for path in XLSX_CANDIDATES:
        if os.path.isfile(path):
            return path
    return None

KNOWN_SOURCE_FRACTIONS = {
    ('Wallet_Users', '2078 Mangsir'),
    ('Mobile_Banking_Count', '2081 Asoj'),
    ('QR_Count', '2081 Asoj'),
}

BS_TO_AD_MONTH = {
    'Baisakh': 4, 'Jestha': 5, 'Asar': 6, 'Saun': 7,
    'Bhadau': 8, 'Asoj': 9, 'Kartik': 10, 'Mangsir': 11,
    'Push': 12, 'Magh': 1, 'Falgun': 2, 'Chaitra': 3
}

def bs_to_ad(year_bs, month_name):
    ad_month = BS_TO_AD_MONTH[month_name]
    ad_year = year_bs - 57 if ad_month >= 4 else year_bs - 56
    return ad_year, ad_month

def get_monthly_col_indices(header_row):
    """Return 0-indexed column indices for monthly data (skip Particulars, FY totals, None)."""
    indices = []
    for i, cell in enumerate(header_row):
        v = cell.value
        if v is None:
            continue
        text = str(v).strip()
        if text.lower().startswith('particular') or 'FY' in str(v):
            continue
        indices.append(i)
    return indices

XLSX_PATH = resolve_xlsx()
if XLSX_PATH is None:
    raise SystemExit(
        'No NRB DataPSD workbook found. Restore sources/NRB_PSD_Sep2025_REAL.xlsx '
        'or add its path to XLSX_CANDIDATES. Refusing to overwrite the corrected CSV.'
    )

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
if 'DataPSD' not in wb.sheetnames:
    raise SystemExit(f'{XLSX_PATH} has no DataPSD sheet. Sheets: {wb.sheetnames}')
ws = wb['DataPSD']
print(f'Source workbook: {XLSX_PATH}')

# === Build canonical month list from Section A (no FY totals) ===
a_header = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
months = []
for i, v in enumerate(a_header):
    if not v:
        continue
    text = str(v).strip()
    if text.lower().startswith('particular'):
        continue
    parts = text.split()
    if len(parts) < 2:
        continue
    months.append({'col_a': i, 'bs_year': int(parts[0]), 'bs_month': parts[1]})

print(f'Months: {len(months)} ({months[0]["bs_year"]} {months[0]["bs_month"]} to {months[-1]["bs_year"]} {months[-1]["bs_month"]})')

# === Get monthly column indices for B1 and B2 ===
b1_monthly = get_monthly_col_indices([ws.cell(row=25, column=c) for c in range(1, ws.max_column + 1)])
b2_monthly = get_monthly_col_indices([ws.cell(row=48, column=c) for c in range(1, ws.max_column + 1)])

print(f'B1 monthly cols: {len(b1_monthly)} (FY totals at: {[i for i in range(1,68) if i-1 not in b1_monthly and i>1]})')
print(f'B2 monthly cols: {len(b2_monthly)}')
assert len(b1_monthly) == len(months), f"B1 ({len(b1_monthly)}) != months ({len(months)})"
assert len(b2_monthly) == len(months), f"B2 ({len(b2_monthly)}) != months ({len(months)})"

# === Row definitions ===
A_FIELDS = [
    ('PSP_Agents', 7), ('Wallet_Users', 8), ('ATM_Machines', 9),
    ('Debit_Cards', 10), ('Credit_Cards', 11), ('Prepaid_Cards', 12),
    ('Mobile_Banking_Customers', 13), ('Internet_Banking_Customers', 14),
    ('Branchless_Banking', 15), ('RTGS_Participants', 16),
    ('ConnectIPS_Users', 17), ('ECC_Members', 18), ('IPS_Members', 19),
]

B1_FIELDS = [
    ('RTGS_Count', 26), ('ATM_Cash_Withdrawal_Count', 27), ('ECC_Count', 28),
    ('IPS_Count', 29), ('ConnectIPS_Count', 30), ('Debit_Card_Count', 31),
    ('Credit_Card_Count', 32), ('Prepaid_Card_Count', 33),
    ('Internet_Banking_Count', 34), ('Mobile_Banking_Count', 35),
    ('Branchless_Banking_Count', 36), ('Wallet_Count', 37), ('QR_Count', 38),
    ('POS_Count', 39), ('Ecommerce_Count', 40), ('Faster_Payment_Count', 41),
    ('Cross_Border_QR_Count', 42), ('Other_Retail_Count', 43),
]

B2_FIELDS = [
    ('RTGS_Value', 49), ('ATM_Cash_Withdrawal_Value', 50), ('ECC_Value', 51),
    ('IPS_Value', 52), ('ConnectIPS_Value', 53), ('Debit_Card_Value', 54),
    ('Credit_Card_Value', 55), ('Prepaid_Card_Value', 56),
    ('Internet_Banking_Value', 57), ('Mobile_Banking_Value', 58),
    ('Branchless_Banking_Value', 59), ('Wallet_Value', 60), ('QR_Value', 61),
    ('POS_Value', 62), ('Ecommerce_Value', 63), ('Faster_Payment_Value', 64),
    ('Cross_Border_QR_Value', 65), ('Other_Retail_Value', 66),
]

CSV_COLUMNS = (['date_bs', 'date_ad'] + 
               [f for f, _ in A_FIELDS] + 
               [f for f, _ in B1_FIELDS] + 
               [f for f, _ in B2_FIELDS])

# === Extract ===
fractional_in_source = []
rows_written = 0

with open(OUTPUT_PATH, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(CSV_COLUMNS)

    for idx, m in enumerate(months):
        ad_year, ad_month = bs_to_ad(m['bs_year'], m['bs_month'])
        date_bs = f'{m["bs_year"]} {m["bs_month"]}'
        date_ad = f'{ad_year}-{ad_month:02d}'
        row = [date_bs, date_ad]

        # Section A: col_a is 0-indexed, openpyxl wants 1-indexed
        for field_name, excel_row in A_FIELDS:
            val = ws.cell(row=excel_row, column=m['col_a'] + 1).value
            if val is not None:
                if isinstance(val, float) and val != int(val):
                    fractional_in_source.append((field_name, date_bs, val))
                    row.append(val)  # Keep the actual fractional value
                else:
                    row.append(int(val))
            else:
                row.append('')

        # Section B1: use b1_monthly[idx] to get correct column
        b1_col = b1_monthly[idx]
        for field_name, excel_row in B1_FIELDS:
            val = ws.cell(row=excel_row, column=b1_col + 1).value
            if val is not None:
                if isinstance(val, float) and val != int(val):
                    fractional_in_source.append((field_name, date_bs, val))
                    row.append(val)
                else:
                    row.append(int(val))
            else:
                row.append('')

        # Section B2: use b2_monthly[idx]
        b2_col = b2_monthly[idx]
        for field_name, excel_row in B2_FIELDS:
            val = ws.cell(row=excel_row, column=b2_col + 1).value
            if val is not None:
                row.append(val)
            else:
                row.append('')

        writer.writerow(row)
        rows_written += 1

print(f'\nWrote {rows_written} rows to {OUTPUT_PATH}')
print(f'Columns: {len(CSV_COLUMNS)}')

# Report fractional values from source
print(f'\n=== FRACTIONAL VALUES PRESENT IN NRB SOURCE XLSX ===')
if fractional_in_source:
    for field, date, val in fractional_in_source:
        print(f'  {field} @ {date}: {val}')
else:
    print('  None found')

# Validation: verify no FY totals leaked
print(f'\n=== VALIDATION ===')
errors = 0
with open(OUTPUT_PATH, 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        for count_field in ['Mobile_Banking_Count', 'Wallet_Count', 'QR_Count', 
                           'RTGS_Count', 'Debit_Card_Count']:
            v = row.get(count_field, '')
            if v and v != '':
                try:
                    fv = float(v)
                    if fv != int(fv):
                        key = (count_field, row['date_bs'])
                        if key in KNOWN_SOURCE_FRACTIONS:
                            print(f'  KNOWN NRB FRACTION: {count_field} @ {row["date_bs"]} = {v}')
                        else:
                            print(f'  UNEXPECTED FRACTIONAL COUNT: {count_field} @ {row["date_bs"]} = {v}')
                            errors += 1
                except ValueError:
                    pass

if errors == 0:
    print('  Count fields are integers except the documented NRB source fractions.')
else:
    print(f'  Validation failures: {errors}')
print(f'\nDone.')
